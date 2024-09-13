import config
import telebot
from telebot import types
from telebot.types import WebAppInfo
import pandas as pd
import sqlite3
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

bot = telebot.TeleBot(config.token)


@bot.message_handler(commands=['commands'])
def handle_command(message):
    keyboard = types.InlineKeyboardMarkup()

    keyboard.add(types.InlineKeyboardButton("Add product", url="http://195.189.227.67:8080"))
    keyboard.add(types.InlineKeyboardButton("Delete product", url="http://195.189.227.67:8080/delete"))
    keyboard.add(types.InlineKeyboardButton("Show product", url="http://195.189.227.67:8080/show"))
    keyboard.add(types.InlineKeyboardButton("Change product", url="http://195.189.227.67:8080/change"))
    keyboard.add(types.InlineKeyboardButton("Add Order", url="http://195.189.227.67:8080/form"))
    keyboard.add(types.InlineKeyboardButton("Orders List", url="http://195.189.227.67:8080/orders"))
    keyboard.add(types.InlineKeyboardButton("Other commands", callback_data="other"))

    bot.send_message(message.from_user.id, "Database commands:", reply_markup=keyboard)

@bot.callback_query_handler(func=lambda call: True)
def callback_query(call):
    if call.data == "other":

        text = """/clients - для отримання клієнтів
/profit_sum YY-mm-dd YY-mm-dd - для цієї команди ввести 2 дати в такому форматі для профіту загального прибутку за період
/profit_all YY-mm-dd YY-mm-dd - для прибутку для кожного продукту за період
"""

        bot.send_message(call.message.chat.id, text)

@bot.message_handler(commands=['clients'])
def handle_command(message):
    conn = sqlite3.connect('web/clients.db')
    query = "SELECT * FROM Clients"

    df = pd.read_sql_query(query, conn)
    
    excel_file_path = 'clients.xlsx'
    df.to_excel(excel_file_path, index=False)

    f = open("clients.xlsx", "rb")
    bot.send_document(message.from_user.id, f)

    f.close()
    conn.close()
    os.remove("clients.xlsx")


@bot.message_handler(commands=['profit_sum'])
def handle_command(message):

    date_from = message.text.split(' ')[1]
    date_to = message.text.split(' ')[2]

    conn = sqlite3.connect('web/profit.db')
    query = f"""
        WITH total AS (
        SELECT id,
            SUM((sell_price-buy_price)*quantity) AS margin
        FROM (
            SELECT * FROM Profit
            WHERE date >= DATE('{date_from}') AND date <= DATE('{date_to}')
        )
        GROUP BY id
        ),
        types AS (
        SELECT id,
               type
        FROM Profit
        )
            
        SELECT DISTINCT types.id,
               margin,
               total_sum,
               types.type
        FROM total
        JOIN
        (SELECT SUM(margin) AS total_sum
        FROM total)
        JOIN types
        ON types.id = total.id

        """

    df = pd.read_sql_query(query, conn)
    
    excel_file_path = 'profit.xlsx'
    df.to_excel(excel_file_path, index=False)


    # wb = load_workbook('existing_file.xlsx')
    # ws = wb.active

    f = open("profit.xlsx", "rb")
    bot.send_document(message.from_user.id, f)

    f.close()
    conn.close()
    os.remove("profit.xlsx")

@bot.message_handler(commands=['profit_all'])
def handle_command(message):

    date_from = message.text.split(' ')[1]
    date_to = message.text.split(' ')[2]
    
    conn = sqlite3.connect('web/profit.db')
    query = f"""
    SELECT * 
    FROM Profit
    WHERE date >= DATE('{date_from}') AND date <= DATE('{date_to}')
    """

    df = pd.read_sql_query(query, conn)
    ids = df['id'].tolist()
    
    excel_file_path = 'profit.xlsx'
    df.to_excel(excel_file_path, index=False)

    wb = load_workbook('profit.xlsx')
    ws = wb.active

    for idx, i in enumerate(ids):
        img = Image(f"web/static/images/{i}.jpg")
        img.width = 100
        img.height = 100

        ws.add_image(img, f'H{idx+2}')
    
    for i in range(1, len(ids)+1):
        ws.row_dimensions[i].height = 100


    wb.save('profit.xlsx')

    f = open("profit.xlsx", "rb")
    
    bot.send_document(message.from_user.id, f)

    f.close()
    conn.close()
    os.remove("profit.xlsx")


if __name__ == '__main__':
    bot.infinity_polling()
